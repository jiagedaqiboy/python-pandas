import time
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill # 样式颜色
openurl = '2021年图纸更新记录.xlsx' # 保存路径
# 显示当前年月
nows= time.localtime(time.time())
month = time.strftime("%Y年%m月",nows)

# 打开图纸自动更新表，选中图纸日期
old = pd.read_excel(r'图纸自动更新.xlsx', sheet_name='图纸刷选', header=[0,1])
troc_Time = old.iloc[:, 4]
x55_Time = old.iloc[:, 8]
vw380_Time = old.iloc[:, 12]

# 打开图纸更新记录表
wd = pd.read_excel(r'../Data/2021年图纸更新记录.xlsx',sheet_name='T-ROC',header=0,index_col=0)
wd[month] = troc_Time
wd[month] = pd.to_numeric(wd[month])   # 转换数据类型
we = pd.read_excel(r'../Data/2021年图纸更新记录.xlsx',sheet_name='Q2L',header=0,index_col=0)
we[month] = x55_Time
we[month] = pd.to_numeric(we[month])
wf = pd.read_excel(r'../Data/2021年图纸更新记录.xlsx',sheet_name='G8',header=0,index_col=0)
wf[month] = vw380_Time
wf[month] = pd.to_numeric(wf[month])
# 保存
writer = pd.ExcelWriter(openurl)
wd.to_excel(writer, sheet_name='T-ROC')
we.to_excel(writer, sheet_name='Q2L')
wf.to_excel(writer, sheet_name='G8')
writer.save()
time.sleep(3)
# 修改样式，判断与上一个月值是否一直
def reponse(openurl,name):
    wb = openpyxl.load_workbook(openurl)
    worksheet = wb[name]
    worksheet.column_dimensions['B'].width = 30
    worksheet.column_dimensions['C'].width = 15
    worksheet.column_dimensions['P'].width = 35
    fill = PatternFill('solid', start_color='FF0000')  # 设置填充颜色为 红色
    for a in range(2, 26):
        for b in range(4, 16):
            c = worksheet.cell(row=a, column=b).value
            d = worksheet.cell(row=a, column=b + 1).value
            if c != d:
                worksheet.cell(row=a, column=b + 1).fill = fill
    wb.save('2021年图纸更新记录.xlsx')

c1 = reponse(openurl,'T-ROC')
c2 = reponse(openurl,'Q2L')
c3 = reponse(openurl,'G8')

